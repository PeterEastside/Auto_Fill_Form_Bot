from openpyxl import load_workbook
import os
import subprocess
import time
import pyautogui as pg
import pyperclip

path = os.getcwd()

print(path)
excelpath = os.path.join(path,"databot.xlsx")

excelfile = load_workbook(excelpath)

sheet = excelfile.active
#print(sheet["A1"].value)

count = len(sheet["A"])


print(count)

machines = [sheet.cell(row=i,column=1).value for i in range(2,count+1)]
problems = [sheet.cell(row=i,column=2).value for i in range(2,count+1)]
reporters = [sheet.cell(row=i,column=3).value for i in range(2,count+1)]
print(machines)
print(problems)
print(reporters)

record = list(zip(machines,problems,reporters))
print(record)
#print(sheet.cell(row=1,column=1).value)

pathprogram = "C:\\Users\\User\\Desktop\\Maintenance-Custom-main\\maintenance.exe"
subprocess.Popen(pathprogram)

time.sleep(3)
pg.press("tab")

for a,b,c in record:
    pyperclip.copy(a)
    time.sleep(0.2)
    pg.hotkey("ctrl","v")
    pg.press("tab")
    
    pyperclip.copy(b)
    time.sleep(0.2)
    pg.hotkey("ctrl","v")
    pg.press("tab")
    
    pyperclip.copy(c)
    time.sleep(0.2)
    pg.hotkey("ctrl","v")
    pg.press("tab")
    
    pg.press("space")
    pg.press("tab",presses=2)
    time.sleep(2)